"""
注文データCSVを1回読み込み、以下の2つの処理をまとめて行うスクリプト。

  1. データ整形(元データ／振り分け／まとめ)をExcelファイルとして出力
  2. 注文者ごとの見積書をPDFとして出力

PyInstallerでexe化し、Excel(.xlsm)のボタンから呼び出すことを想定している。
PDF出力にはExcelの実機(win32com)を使うため、実行するPCにMicrosoft Excelが
インストールされている必要がある。
"""

import os
import sys
import time
import traceback
import tkinter as tk
from datetime import date
from tkinter import filedialog, messagebox

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

try:
    import win32com.client
except ImportError:
    win32com = None


# ============================================================
# 共通設定: 元データ(CSV)の列位置(0始まりのインデックス)
# ============================================================
ORDER_DATE_COL = 2      # 3列目: 注文日
PRODUCT_NAME_COL = 4    # 5列目: 商品名
PRODUCT_CODE_COL = 5    # 6列目: 商品番号
QUANTITY_COL = 6        # 7列目: 個数
UNIT_PRICE_COL = 7      # 8列目: 単価
LAST_NAME_COL = 9       # 10列目: 注文者名字
FIRST_NAME_COL = 10     # 11列目: 注文者名前

# 整形データの出力シート名
SHEET_NAME_RAW = "元データ"
SHEET_NAME_FORMATTED_1 = "振り分け"
SHEET_NAME_FORMATTED_2 = "まとめ"

# 見積書テンプレート側のセル位置
CELL_CUSTOMER_NAME = "A3"
CELL_SUBJECT = "C6"
CELL_QUOTE_NO = "N3"
CELL_QUOTE_DATE = "N4"
DETAIL_START_ROW = 18
DETAIL_END_ROW = 29
DETAIL_MAX_ROWS = DETAIL_END_ROW - DETAIL_START_ROW + 1  # 12
COL_PRODUCT_CODE = "B"
COL_QUANTITY = "J"
COL_UNIT_PRICE = "L"

# exeと同じフォルダに置く見積書テンプレートのファイル名
TEMPLATE_FILE_NAME = "Invoice_Template.xlsx"


def resource_path(filename: str) -> str:
    """
    exe化した際、実行ファイルと同じフォルダにあるファイルのパスを返す。
    (通常のpython実行時はこのスクリプトと同じフォルダを見る)
    """
    if getattr(sys, "frozen", False):
        base_dir = os.path.dirname(sys.executable)
    else:
        base_dir = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base_dir, filename)


def select_csv_file() -> str:
    root = tk.Tk()
    root.withdraw()
    return filedialog.askopenfilename(
        title="注文データCSVを選択してください",
        filetypes=[("CSVファイル", "*.csv"), ("すべてのファイル", "*.*")],
    )


def load_orders(csv_path: str) -> pd.DataFrame:
    return pd.read_csv(csv_path, encoding="cp932")


# ============================================================
# ① データ整形(元データ／振り分け／まとめ)
# ============================================================

def format_dataframe_1(df: pd.DataFrame) -> pd.DataFrame:
    """「振り分け」: 個数の数だけ行を複製し、複製後の個数はすべて1にする"""
    qty_col = df.columns[QUANTITY_COL]
    repeat_counts = (
        pd.to_numeric(df[qty_col], errors="coerce")
        .fillna(1).astype(int).clip(lower=1)
    )
    df_expanded = df.loc[df.index.repeat(repeat_counts)].reset_index(drop=True)
    df_expanded[qty_col] = 1
    return df_expanded


def format_dataframe_2(df: pd.DataFrame):
    """「まとめ」: 商品番号ごとに集計し、商品名・単価の食い違いは警告として返す"""
    name_col = df.columns[PRODUCT_NAME_COL]
    code_col = df.columns[PRODUCT_CODE_COL]
    qty_col = df.columns[QUANTITY_COL]
    price_col = df.columns[UNIT_PRICE_COL]

    df = df.copy()
    df[qty_col] = pd.to_numeric(df[qty_col], errors="coerce").fillna(0)
    df[price_col] = pd.to_numeric(df[price_col], errors="coerce").fillna(0)

    warnings_list = []
    name_nunique = df.groupby(code_col)[name_col].nunique()
    price_nunique = df.groupby(code_col)[price_col].nunique()

    for code in name_nunique[name_nunique > 1].index:
        names = df.loc[df[code_col] == code, name_col].unique().tolist()
        msg = f"商品番号「{code}」で商品名が一致していません: {names}"
        warnings_list.append(msg)
        print(f"⚠ 警告: {msg}")

    for code in price_nunique[price_nunique > 1].index:
        prices = df.loc[df[code_col] == code, price_col].unique().tolist()
        msg = f"商品番号「{code}」で単価が一致していません: {prices}"
        warnings_list.append(msg)
        print(f"⚠ 警告: {msg}")

    grouped = df.groupby(code_col, as_index=False).agg({
        name_col: "first", qty_col: "sum", price_col: "first",
    })
    grouped = grouped[[name_col, code_col, qty_col, price_col]]
    grouped["合計金額"] = grouped[qty_col] * grouped[price_col]
    return grouped, warnings_list


def format_worksheet(worksheet, df: pd.DataFrame):
    """列幅の自動調整とヘッダーの太字・背景色を設定する"""
    CONTENT_PADDING = 2
    HEADER_PADDING = 4
    for i, col in enumerate(df.columns, start=1):
        if len(df) > 0:
            content_max_len = df[col].map(lambda x: len(str(x))).max()
        else:
            content_max_len = 0
        max_len = max(content_max_len + CONTENT_PADDING, len(str(col)) + HEADER_PADDING)
        worksheet.column_dimensions[get_column_letter(i)].width = max_len

    for cell in worksheet[1]:
        cell.font = Font(bold=True)
    for cell in worksheet[1]:
        cell.fill = PatternFill(start_color="D9E1F2", fill_type="solid")


def save_formatted_excel(df_raw, df_formatted_1, df_formatted_2, csv_path) -> str:
    """元データ・振り分け・まとめの3シートを持つExcelファイルを保存する"""
    base, _ = os.path.splitext(csv_path)
    excel_path = base + ".xlsx"

    sheets = {
        SHEET_NAME_RAW: df_raw,
        SHEET_NAME_FORMATTED_1: df_formatted_1,
        SHEET_NAME_FORMATTED_2: df_formatted_2,
    }
    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        for sheet_name, df in sheets.items():
            df.to_excel(writer, index=False, sheet_name=sheet_name)
            worksheet = writer.sheets[sheet_name]
            format_worksheet(worksheet, df)

    return excel_path


# ============================================================
# ② 見積書(PDF)作成
# ============================================================

def build_customer_groups(df: pd.DataFrame):
    """注文者(姓+名)ごとにグループ化し、商品番号ごとに集計した明細を返す"""
    last_col = df.columns[LAST_NAME_COL]
    first_col = df.columns[FIRST_NAME_COL]
    date_col = df.columns[ORDER_DATE_COL]
    name_col = df.columns[PRODUCT_NAME_COL]
    code_col = df.columns[PRODUCT_CODE_COL]
    qty_col = df.columns[QUANTITY_COL]
    price_col = df.columns[UNIT_PRICE_COL]

    df = df.copy()
    df["_customer_name"] = df[last_col].astype(str) + df[first_col].astype(str)
    df[qty_col] = pd.to_numeric(df[qty_col], errors="coerce").fillna(0)
    df[price_col] = pd.to_numeric(df[price_col], errors="coerce").fillna(0)

    groups = []
    for customer_name, customer_df in df.groupby("_customer_name", sort=False):
        # 商品番号ごとに集計(商品名・単価は先頭の値を採用、個数は合計)
        detail = customer_df.groupby(code_col, as_index=False).agg({
            name_col: "first", qty_col: "sum", price_col: "first",
        })
        # 見積書には商品名を書き込むため、商品名・個数・単価の順にする
        detail = detail[[name_col, qty_col, price_col]]
        subject_date = pd.to_datetime(customer_df[date_col], errors="coerce").min()
        groups.append((customer_name, subject_date, detail))

    return groups


def chunk_detail(detail: pd.DataFrame, chunk_size: int = DETAIL_MAX_ROWS):
    """12種類を超える場合に備え、明細をchunk_size行ごとに分割する"""
    for i in range(0, len(detail), chunk_size):
        yield detail.iloc[i:i + chunk_size]


def to_number_for_cell(value):
    try:
        if float(value).is_integer():
            return int(value)
    except (TypeError, ValueError):
        pass
    return value


def fill_quote(template_path, customer_name, subject_text, detail_chunk, quote_no, quote_date_text):
    """テンプレートに1件分の見積書データを書き込んだWorkbookを返す"""
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active
    ws[CELL_CUSTOMER_NAME] = customer_name
    ws[CELL_SUBJECT] = subject_text
    ws[CELL_QUOTE_NO] = quote_no
    ws[CELL_QUOTE_DATE] = quote_date_text

    # O列(金額)は数式(=J×L)が既に入っているため触らない
    for offset, (_, row) in enumerate(detail_chunk.iterrows()):
        r = DETAIL_START_ROW + offset
        ws[f"A{r}"] = offset + 1            # 連番(1, 2, 3...)
        ws[f"{COL_PRODUCT_CODE}{r}"] = row.iloc[0]  # 商品名
        ws[f"{COL_QUANTITY}{r}"] = to_number_for_cell(row.iloc[1])
        ws[f"{COL_UNIT_PRICE}{r}"] = to_number_for_cell(row.iloc[2])

    return wb


def export_excel_to_pdf(excel_app, xlsx_path: str, pdf_path: str, retries: int = 2):
    """
    Excel実機(win32com)でxlsxファイルを開き、PDFとして書き出す。
    この時点で数式(O列など)が計算された状態でPDF化される。

    パスは絶対パスに変換してから渡す(相対パスだとExcel側で
    「Document not saved」エラーになることがあるため)。
    一時的なファイルロックなどで失敗することがあるため、
    数回リトライする。
    """
    xlsx_path = os.path.abspath(xlsx_path)
    pdf_path = os.path.abspath(pdf_path)

    last_error = None
    for attempt in range(1, retries + 1):
        wb = None
        try:
            wb = excel_app.Workbooks.Open(xlsx_path, UpdateLinks=0, ReadOnly=False)
            # 明示的に保存しておく(「保存されていません」エラー対策)
            wb.Save()
            wb.ExportAsFixedFormat(0, pdf_path)  # 0 = xlTypePDF
            return
        except Exception as e:
            last_error = e
            print(f"PDF出力に失敗しました(試行{attempt}/{retries}): {e}")
            time.sleep(1.5)
        finally:
            if wb is not None:
                wb.Close(SaveChanges=False)

    raise last_error


def sanitize_filename(text: str) -> str:
    """ファイル名に使えない文字を全角に置き換える"""
    invalid_chars = {
        "\\": "￥", "/": "／", ":": "：", "*": "＊",
        "?": "？", '"': "”", "<": "＜", ">": "＞", "|": "｜",
    }
    for char, replacement in invalid_chars.items():
        text = text.replace(char, replacement)
    return text


def make_quote_pdfs(df: pd.DataFrame, template_path: str, output_dir: str) -> list:
    """注文者ごとに見積書PDFを作成し、保存したパスのリストを返す"""
    if win32com is None:
        raise RuntimeError(
            "pywin32がインストールされていないため、PDF出力ができません。"
            "「pip install pywin32」を実行してください。"
        )

    groups = build_customer_groups(df)

    today = date.today()
    quote_date_text = f"{today.year}年{today.month}月{today.day}日"
    date_prefix = today.strftime("%Y%m%d")
    seq = 1
    saved_pdfs = []

    excel_app = win32com.client.Dispatch("Excel.Application")
    excel_app.Visible = False
    excel_app.DisplayAlerts = False
    excel_app.AskToUpdateLinks = False

    output_dir = os.path.abspath(output_dir)

    try:
        for customer_name, subject_date, detail in groups:
            chunks = list(chunk_detail(detail))
            multi_page = len(chunks) > 1  # 12種類を超える場合は複数ファイルに分割

            # 件名: 「yyyy年m月d日ご注文分」(このお客様全体で共通)
            if pd.notna(subject_date):
                subject_text = f"{subject_date.year}年{subject_date.month}月{subject_date.day}日ご注文分"
            else:
                subject_text = "ご注文分"

            for page_index, chunk in enumerate(chunks, start=1):
                quote_no = f"{date_prefix}{seq:04d}"
                seq += 1

                wb = fill_quote(
                    template_path, customer_name, subject_text, chunk,
                    quote_no, quote_date_text,
                )

                # ファイル名: 「見積書_氏名様_件名」(複数ページの場合は末尾に連番)
                suffix = f"_{page_index}" if multi_page else ""
                base_name = sanitize_filename(f"見積書_{customer_name}様_{subject_text}{suffix}")
                temp_xlsx_path = os.path.join(output_dir, base_name + ".xlsx")
                pdf_path = os.path.join(output_dir, base_name + ".pdf")

                wb.save(temp_xlsx_path)
                export_excel_to_pdf(excel_app, temp_xlsx_path, pdf_path)

                # 中間ファイル(xlsx)は不要なので削除する
                os.remove(temp_xlsx_path)

                saved_pdfs.append(pdf_path)
    finally:
        excel_app.Quit()

    return saved_pdfs


# ============================================================
# メイン処理: CSVを1回読み込み、整形とPDF見積書の両方を行う
# ============================================================

def process_orders():
    csv_path = select_csv_file()
    if not csv_path:
        print("CSVファイルが選択されませんでした。処理を終了します。")
        return

    template_path = resource_path(TEMPLATE_FILE_NAME)
    if not os.path.exists(template_path):
        messagebox.showerror(
            "エラー",
            f"見積書テンプレートが見つかりません:\n{template_path}\n"
            "実行ファイルと同じフォルダにテンプレートを置いてください。",
        )
        return

    try:
        df_raw = load_orders(csv_path)

        # ① データ整形
        df_formatted_1 = format_dataframe_1(df_raw)
        df_formatted_2, warnings_list = format_dataframe_2(df_raw)
        excel_path = save_formatted_excel(df_raw, df_formatted_1, df_formatted_2, csv_path)

        # ② 見積書(PDF)作成
        pdf_output_dir = os.path.join(os.path.dirname(csv_path), "見積書")
        os.makedirs(pdf_output_dir, exist_ok=True)
        saved_pdfs = make_quote_pdfs(df_raw, template_path, pdf_output_dir)

        result_msg = (
            f"データ整形: {excel_path}\n"
            f"見積書PDF: {len(saved_pdfs)}件\n({pdf_output_dir})"
        )
        print(result_msg)

        if warnings_list:
            warning_text = "\n".join(f"・{w}" for w in warnings_list)
            messagebox.showwarning("警告", f"以下の不整合がありました:\n\n{warning_text}")

        messagebox.showinfo("完了", result_msg)

    except Exception as e:
        print("エラーが発生しました:")
        traceback.print_exc()
        messagebox.showerror("エラー", f"処理中にエラーが発生しました:\n{e}")


if __name__ == "__main__":
    process_orders()